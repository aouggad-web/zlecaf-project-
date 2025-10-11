#!/usr/bin/env python3
"""
Script pour créer le fichier Excel de validation des données ZLECAf
"""
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import csv

def create_validation_file():
    # Lire le fichier CSV existant
    df = pd.read_csv('/app/ZLECAF_54_PAYS_DONNEES_COMPLETES.csv')
    
    # Ajouter des colonnes pour la validation
    df['VALIDATION_STATUT'] = df['Notes_Validation']
    df['CORRECTIONS_PIB'] = ''
    df['CORRECTIONS_POPULATION'] = ''
    df['CORRECTIONS_IDH'] = ''
    df['CORRECTIONS_SECTEURS'] = ''
    df['COMMENTAIRES_GENERAUX'] = ''
    df['SOURCES_SUPPLEMENTAIRES'] = ''
    df['VALIDE_PAR'] = ''
    df['DATE_VALIDATION'] = ''
    
    # Réorganiser les colonnes pour la validation
    columns_order = [
        'Pays', 'Code_ISO', 'VALIDATION_STATUT',
        'PIB_2024_Mds_USD', 'CORRECTIONS_PIB',
        'Population_2024_M', 'CORRECTIONS_POPULATION', 
        'PIB_par_habitant_USD',
        'IDH_2024', 'CORRECTIONS_IDH', 'Rang_Afrique_IDH',
        'Croissance_2024_Pct',
        'Secteur_1', 'Part_Secteur_1_Pct',
        'Secteur_2', 'Part_Secteur_2_Pct', 
        'Secteur_3', 'Part_Secteur_3_Pct',
        'CORRECTIONS_SECTEURS',
        'Sources_Principales', 'SOURCES_SUPPLEMENTAIRES',
        'COMMENTAIRES_GENERAUX',
        'VALIDE_PAR', 'DATE_VALIDATION'
    ]
    
    df_validation = df[columns_order]
    
    # Créer le fichier Excel avec formatage
    with pd.ExcelWriter('/app/ZLECAF_DONNEES_VALIDATION.xlsx', engine='openpyxl') as writer:
        df_validation.to_excel(writer, sheet_name='Données à Valider', index=False)
        
        # Obtenir la feuille de travail
        ws = writer.sheets['Données à Valider']
        
        # Définir les styles - African-inspired color scheme
        header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')  # African green
        header_font = Font(color='FFFFFF', bold=True)
        validation_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')  # Light gold/amber
        error_fill = PatternFill(start_color='FFCCBC', end_color='FFCCBC', fill_type='solid')  # Light orange
        complete_fill = PatternFill(start_color='A5D6A7', end_color='A5D6A7', fill_type='solid')  # Light green
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Formatter l'en-tête
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Ajuster la largeur des colonnes
        column_widths = {
            'A': 20, 'B': 8, 'C': 20, 'D': 15, 'E': 20,
            'F': 15, 'G': 25, 'H': 15, 'I': 12, 'J': 20,
            'K': 15, 'L': 15, 'M': 15, 'N': 12, 'O': 15,
            'P': 12, 'Q': 15, 'R': 12, 'S': 25, 'T': 30,
            'U': 30, 'V': 40, 'W': 15, 'X': 15
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Colorer les lignes selon le statut de validation
        for row in range(2, ws.max_row + 1):
            status = ws[f'C{row}'].value
            
            if status and 'À compléter' in str(status):
                fill = error_fill
            elif status and 'à valider' in str(status):
                fill = validation_fill
            elif status and 'vérifiées' in str(status):
                fill = complete_fill
            else:
                fill = validation_fill
            
            # Appliquer la couleur à toute la ligne
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
        
        # Figer la première ligne
        ws.freeze_panes = 'A2'
        
        # Ajouter une feuille d'instructions
        instructions = [
            ["INSTRUCTIONS POUR LA VALIDATION DES DONNÉES ZLECAF"],
            [""],
            ["LÉGENDE DES COULEURS:"],
            ["🟢 VERT: Données vérifiées et validées"],
            ["🟡 JAUNE: Données à valider - nécessitent vérification"],
            ["🔴 ROSE: Données incomplètes - à compléter"],
            [""],
            ["COLONNES DE VALIDATION:"],
            ["• VALIDATION_STATUT: Statut actuel de la donnée"],
            ["• CORRECTIONS_PIB: Vos corrections pour le PIB si nécessaire"],
            ["• CORRECTIONS_POPULATION: Vos corrections pour la population"],
            ["• CORRECTIONS_IDH: Vos corrections pour l'IDH"], 
            ["• CORRECTIONS_SECTEURS: Corrections pour les secteurs économiques"],
            ["• SOURCES_SUPPLEMENTAIRES: Ajoutez vos sources de données"],
            ["• COMMENTAIRES_GENERAUX: Vos remarques et observations"],
            ["• VALIDE_PAR: Votre nom/initiales après validation"],
            ["• DATE_VALIDATION: Date de votre validation"],
            [""],
            ["PRIORITÉS DE VALIDATION:"],
            ["1. Pays roses (À compléter) - PRIORITÉ HAUTE"],
            ["2. Pays jaunes (À valider) - PRIORITÉ MOYENNE"], 
            ["3. Pays verts (Vérifiées) - Contrôle qualité"],
            [""],
            ["SOURCES RECOMMANDÉES:"],
            ["• PNUD: undp.org (IDH)"],
            ["• Banque Mondiale: worldbank.org (PIB, Population)"],
            ["• BAD: afdb.org (Données africaines)"],
            ["• FMI: imf.org (Projections économiques)"],
            ["• Instituts nationaux de statistiques"],
            [""],
            ["MÉTHODE:"],
            ["1. Vérifiez chaque donnée avec au moins 2 sources"],
            ["2. Notez vos corrections dans les colonnes dédiées"],
            ["3. Ajoutez vos commentaires et sources"],
            ["4. Validez avec vos initiales et la date"],
        ]
        
        instructions_df = pd.DataFrame(instructions, columns=['Instructions'])
        instructions_df.to_excel(writer, sheet_name='Instructions', index=False, header=False)
        
        # Formatter la feuille d'instructions
        ws_inst = writer.sheets['Instructions']
        ws_inst.column_dimensions['A'].width = 80
        
        # Titre en gras
        ws_inst['A1'].font = Font(bold=True, size=14)
        ws_inst['A3'].font = Font(bold=True)
        ws_inst['A8'].font = Font(bold=True)
        ws_inst['A19'].font = Font(bold=True)
        ws_inst['A25'].font = Font(bold=True)
        ws_inst['A32'].font = Font(bold=True)
    
    print("✅ Fichier de validation créé: /app/ZLECAF_DONNEES_VALIDATION.xlsx")
    
    # Créer aussi un résumé du statut actuel
    status_summary = df['Notes_Validation'].value_counts()
    print("\n📊 RÉSUMÉ DU STATUT ACTUEL:")
    for status, count in status_summary.items():
        print(f"   {status}: {count} pays")
    
    return df_validation

if __name__ == "__main__":
    create_validation_file()