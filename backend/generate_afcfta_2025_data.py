#!/usr/bin/env python3
"""
Generate AfCFTA Dismantling Matrix and Rates CSVs for 2025-2035
Automates creation of normalized data tables for AfCFTA customs analytics.
"""

import csv
import json
from pathlib import Path
from datetime import datetime
import sys

# Add parent directory to path to import country_data
sys.path.insert(0, str(Path(__file__).parent))
from country_data import REAL_COUNTRY_DATA

# All 54 AfCFTA member countries (ISO3 codes)
ALL_COUNTRIES = sorted(REAL_COUNTRY_DATA.keys())

# Sample HS6 codes for demonstration
# In production, this would be the full HS nomenclature (5000+ codes)
SAMPLE_HS6_CODES = [
    "010121", "010129", "010130", "010190",  # Live animals
    "020110", "020120", "020130",  # Meat
    "030111", "030119", "030192",  # Fish
    "040110", "040120", "040130",  # Dairy
    "070110", "070190", "070200",  # Vegetables
    "080111", "080119", "080300",  # Fruits
    "100110", "100190", "100630",  # Cereals
    "270112", "270119", "270900",  # Mineral fuels
    "280110", "280120", "280300",  # Chemicals
    "390110", "390120", "390130",  # Plastics
    "420100", "420211", "420292",  # Leather goods
    "520100", "520511", "520512",  # Cotton
    "610110", "610120", "610130",  # Knitted apparel
    "620111", "620112", "620113",  # Woven apparel
    "640110", "640192", "640199",  # Footwear
    "720110", "720120", "720150",  # Iron and steel
    "840110", "840120", "840130",  # Machinery
    "850110", "850120", "850131",  # Electrical machinery
    "870110", "870120", "870190",  # Vehicles
]

# AfCFTA dismantling categories
DISMANTLING_CATEGORIES = {
    "A": {"name": "Category A", "description": "Linear reduction to 0% over 5 years", "years": 5},
    "B": {"name": "Category B", "description": "Linear reduction to 0% over 10 years", "years": 10},
    "C": {"name": "Category C", "description": "Linear reduction to 0% over 13 years", "years": 13},
    "Sensitive": {"name": "Sensitive", "description": "Linear reduction to 0% over 15 years", "years": 15},
    "Excluded": {"name": "Excluded", "description": "Excluded from liberalization", "years": None},
}


def generate_dismantling_matrix_wide(output_path: Path):
    """
    Generate wide dismantling matrix: status per HS6 with one column per country.
    Values are blank for data entry (Phase1/2/3, Sensitive, Excluded, Immediate, etc.)
    """
    print("📊 Generating dismantling matrix (wide format)...")
    
    # Headers: HS6 code + all country codes
    headers = ["hs6_code", "hs_description"] + ALL_COUNTRIES
    
    rows = []
    for hs6 in SAMPLE_HS6_CODES:
        # Generic description based on HS chapter
        chapter = hs6[:2]
        description = f"HS Chapter {chapter}"
        row = [hs6, description] + [""] * len(ALL_COUNTRIES)
        rows.append(row)
    
    with open(output_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    
    print(f"   ✅ Created: {output_path.name} ({len(rows)} HS6 codes × {len(ALL_COUNTRIES)} countries)")
    return len(rows)


def generate_rates_by_year_long(output_path: Path, start_year: int = 2025, end_year: int = 2035):
    """
    Generate AfCFTA rates by year (long-form): one row per (country, hs6, year).
    Fields: country, hs6_code, year, mfn_rate_pct, afcfta_rate_pct, schedule, immediate_flag, source_url
    """
    print("📊 Generating rates by year (long format)...")
    
    headers = [
        "country",
        "hs6_code",
        "year",
        "mfn_rate_pct",
        "afcfta_rate_pct",
        "schedule",
        "immediate_flag",
        "source_url"
    ]
    
    rows = []
    # Sample data for a few countries and HS6 codes
    sample_countries = ALL_COUNTRIES[:5]  # First 5 countries
    sample_hs6 = SAMPLE_HS6_CODES[:10]  # First 10 HS6 codes
    
    for country in sample_countries:
        for hs6 in sample_hs6:
            for year in range(start_year, end_year + 1):
                row = [
                    country,
                    hs6,
                    year,
                    "",  # mfn_rate_pct - to be filled
                    "",  # afcfta_rate_pct - to be filled
                    "",  # schedule - to be filled (A, B, C, etc.)
                    "",  # immediate_flag - to be filled (0 or 1)
                    ""   # source_url - to be filled
                ]
                rows.append(row)
    
    with open(output_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    
    print(f"   ✅ Created: {output_path.name} ({len(rows)} rows, {len(sample_countries)} countries × {len(sample_hs6)} HS6 × {end_year - start_year + 1} years)")
    return len(rows)


def generate_immediate_dismantled(output_path: Path, start_year: int = 2025):
    """
    Generate immediate dismantled positions: subset for immediate dismantling cases.
    Fields: country, hs6_code, start_year, initial_mfn_rate_pct, schedule, source_url
    """
    print("📊 Generating immediate dismantled positions...")
    
    headers = [
        "country",
        "hs6_code",
        "start_year",
        "initial_mfn_rate_pct",
        "schedule",
        "notes",
        "source_url"
    ]
    
    rows = []
    # Sample immediate dismantling positions
    sample_countries = ALL_COUNTRIES[:5]  # First 5 countries
    sample_hs6 = SAMPLE_HS6_CODES[:5]  # First 5 HS6 codes
    
    for country in sample_countries:
        for hs6 in sample_hs6:
            row = [
                country,
                hs6,
                start_year,
                "",  # initial_mfn_rate_pct - to be filled
                "Immediate",  # schedule
                "Immediate dismantling - 0% from start",
                ""   # source_url - to be filled
            ]
            rows.append(row)
    
    with open(output_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    
    print(f"   ✅ Created: {output_path.name} ({len(rows)} immediate positions)")
    return len(rows)


def generate_metadata_json(output_path: Path, stats: dict):
    """
    Generate metadata JSON file with generation info and statistics.
    """
    print("📊 Generating metadata...")
    
    metadata = {
        "version": "1.0.0",
        "generated_at": datetime.now().isoformat(),
        "description": "AfCFTA Dismantling Matrix and Rates Data for 2025-2035",
        "source": "ZLECAf Lyra+ Pipeline",
        "period": {
            "start_year": 2025,
            "end_year": 2035,
            "total_years": 11
        },
        "coverage": {
            "total_countries": len(ALL_COUNTRIES),
            "countries": ALL_COUNTRIES,
            "sample_hs6_codes": len(SAMPLE_HS6_CODES),
            "note": "Sample dataset - Production version should include all 5000+ HS6 codes"
        },
        "dismantling_categories": DISMANTLING_CATEGORIES,
        "statistics": stats,
        "data_fields": {
            "dismantling_matrix": {
                "description": "Wide format matrix with HS6 codes as rows and countries as columns",
                "fields": ["hs6_code", "hs_description", "...country_codes"],
                "purpose": "Data entry template for dismantling schedules"
            },
            "rates_by_year": {
                "description": "Long format table with one row per (country, hs6, year)",
                "fields": [
                    "country", "hs6_code", "year", "mfn_rate_pct",
                    "afcfta_rate_pct", "schedule", "immediate_flag", "source_url"
                ],
                "purpose": "Time series data for tariff evolution"
            },
            "immediate_dismantled": {
                "description": "Subset of positions with immediate dismantling (0% from start)",
                "fields": [
                    "country", "hs6_code", "start_year",
                    "initial_mfn_rate_pct", "schedule", "notes", "source_url"
                ],
                "purpose": "Quick reference for immediate liberalization"
            }
        },
        "instructions": {
            "data_entry": "Fill blank fields in dismantling_matrix with schedule types: A, B, C, Sensitive, Excluded, Immediate",
            "rates_entry": "Complete rates_by_year with actual MFN and AfCFTA rates from national schedules",
            "validation": "Ensure immediate_flag=1 when afcfta_rate_pct=0 from start_year"
        }
    }
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(metadata, f, indent=2, ensure_ascii=False)
    
    print(f"   ✅ Created: {output_path.name}")


def generate_excel_workbook(output_path: Path, csv_dir: Path):
    """
    Generate XLSX workbook with multiple sheets.
    Requires openpyxl library.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("   ⚠️  openpyxl not available - skipping XLSX generation")
        print("   💡 Install with: pip install openpyxl")
        return False
    
    print("📊 Generating Excel workbook...")
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Sheet 1: Dismantling Matrix (sample)
    ws_matrix = wb.create_sheet("dismantling_matrix")
    with open(csv_dir / "zlecaf_dismantling_matrix_2025.csv", 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            ws_matrix.append(row)
    
    # Format header row
    for cell in ws_matrix[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    
    # Sheet 2: Rates by Year (sample)
    ws_rates = wb.create_sheet("rates_by_year_sample")
    with open(csv_dir / "zlecaf_afcfta_rates_by_year_2025.csv", 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            ws_rates.append(row)
    
    for cell in ws_rates[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    
    # Sheet 3: Immediate Dismantled (sample)
    ws_immediate = wb.create_sheet("immediate_sample")
    with open(csv_dir / "zlecaf_immediate_dismantled_2025.csv", 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            ws_immediate.append(row)
    
    for cell in ws_immediate[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.font = Font(bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center")
    
    # Sheet 4: Metadata
    ws_metadata = wb.create_sheet("metadata")
    metadata_info = [
        ["AfCFTA Dismantling Matrix and Rates Data 2025-2035", ""],
        ["", ""],
        ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Version", "1.0.0"],
        ["Total Countries", str(len(ALL_COUNTRIES))],
        ["Sample HS6 Codes", str(len(SAMPLE_HS6_CODES))],
        ["Period", "2025-2035"],
        ["", ""],
        ["Sheets", ""],
        ["1. dismantling_matrix", "Wide format: HS6 × Countries"],
        ["2. rates_by_year_sample", "Long format: Country-HS6-Year"],
        ["3. immediate_sample", "Immediate dismantling positions"],
        ["4. metadata", "This information sheet"],
        ["", ""],
        ["Dismantling Categories", ""],
        ["A", "0% in 5 years"],
        ["B", "0% in 10 years"],
        ["C", "0% in 13 years"],
        ["Sensitive", "0% in 15 years"],
        ["Excluded", "No liberalization"],
        ["Immediate", "0% from start"],
    ]
    
    for row in metadata_info:
        ws_metadata.append(row)
    
    # Format metadata sheet
    ws_metadata.column_dimensions['A'].width = 30
    ws_metadata.column_dimensions['B'].width = 40
    for row in [1, 9, 15]:
        ws_metadata.cell(row, 1).font = Font(bold=True, size=14)
        ws_metadata.cell(row, 1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Auto-size columns for data sheets
    for ws in [ws_matrix, ws_rates, ws_immediate]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    wb.save(output_path)
    print(f"   ✅ Created: {output_path.name} (4 sheets)")
    return True


def main():
    """Main execution function"""
    print("=" * 80)
    print("🚀 AFCFTA DISMANTLING MATRIX & RATES DATA GENERATOR 2025-2035")
    print("=" * 80)
    print(f"📅 Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"🌍 Countries: {len(ALL_COUNTRIES)} AfCFTA members")
    print(f"📦 Sample HS6: {len(SAMPLE_HS6_CODES)} codes (production: 5000+)")
    print()
    
    # Determine output directory
    script_dir = Path(__file__).parent.parent
    output_dir = script_dir / "frontend" / "public" / "data"
    output_dir.mkdir(parents=True, exist_ok=True)
    
    print(f"📂 Output directory: {output_dir}")
    print()
    
    # Generate files
    stats = {}
    
    # 1. Wide dismantling matrix
    matrix_path = output_dir / "zlecaf_dismantling_matrix_2025.csv"
    stats["matrix_rows"] = generate_dismantling_matrix_wide(matrix_path)
    
    # 2. Rates by year (long format)
    rates_path = output_dir / "zlecaf_afcfta_rates_by_year_2025.csv"
    stats["rates_rows"] = generate_rates_by_year_long(rates_path)
    
    # 3. Immediate dismantled positions
    immediate_path = output_dir / "zlecaf_immediate_dismantled_2025.csv"
    stats["immediate_rows"] = generate_immediate_dismantled(immediate_path)
    
    # 4. Metadata
    metadata_path = output_dir / "zlecaf_afcfta_2025_metadata.json"
    generate_metadata_json(metadata_path, stats)
    
    # 5. Excel workbook (if openpyxl available)
    excel_path = output_dir / "zlecaf_afcfta_dismantling_2025.xlsx"
    excel_created = generate_excel_workbook(excel_path, output_dir)
    
    print()
    print("=" * 80)
    print("✅ GENERATION COMPLETE")
    print("=" * 80)
    print()
    print("📋 Generated files:")
    for filename in ["zlecaf_dismantling_matrix_2025.csv",
                     "zlecaf_afcfta_rates_by_year_2025.csv",
                     "zlecaf_immediate_dismantled_2025.csv",
                     "zlecaf_afcfta_2025_metadata.json",
                     "zlecaf_afcfta_dismantling_2025.xlsx"]:
        filepath = output_dir / filename
        if filepath.exists():
            size = filepath.stat().st_size
            print(f"   ✓ {filename:50s} ({size:,} bytes)")
    
    print()
    print("📊 Statistics:")
    print(f"   • Dismantling matrix: {stats['matrix_rows']} HS6 codes")
    print(f"   • Rates by year: {stats['rates_rows']} rows")
    print(f"   • Immediate positions: {stats['immediate_rows']} rows")
    print()
    print("💡 Next steps:")
    print("   1. Fill blank fields in dismantling_matrix with schedule types (A, B, C, etc.)")
    print("   2. Complete rates_by_year with actual MFN and AfCFTA rates")
    print("   3. Validate immediate_dismantled against national schedules")
    print("   4. Update with full HS6 nomenclature (5000+ codes) for production")
    print()


if __name__ == "__main__":
    main()
