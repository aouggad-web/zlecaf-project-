#!/usr/bin/env python3
"""
Script de génération de datasets ZLECAf pour Lyra+ Ops
Génère les fichiers de données pour le frontend
"""
import json
import csv
import argparse
import os
from pathlib import Path
from datetime import datetime

# Fichiers à générer
FILES = [
    'zlecaf_tariff_lines_by_country.json',
    'zlecaf_africa_vs_world_tariffs.xlsx',
    'zlecaf_rules_of_origin.json',
    'zlecaf_dismantling_schedule.csv',
    'zlecaf_tariff_origin_phase.json',
]

def generate_demo_data(output_dir: Path):
    """Génère des données de démonstration"""
    print(f"🔄 Génération des données de démonstration...")
    
    # 1. zlecaf_tariff_lines_by_country.json
    tariff_lines = {
        "generated_at": datetime.now().isoformat(),
        "countries": [
            {"code": "DZ", "name": "Algérie", "tariff_lines": 5000},
            {"code": "EG", "name": "Égypte", "tariff_lines": 6500},
            {"code": "KE", "name": "Kenya", "tariff_lines": 5800},
            {"code": "NG", "name": "Nigéria", "tariff_lines": 7200},
            {"code": "ZA", "name": "Afrique du Sud", "tariff_lines": 8100},
        ]
    }
    
    with open(output_dir / FILES[0], 'w', encoding='utf-8') as f:
        json.dump(tariff_lines, f, indent=2, ensure_ascii=False)
    print(f"   ✅ {FILES[0]}")
    
    # 2. zlecaf_africa_vs_world_tariffs.xlsx
    # Pour le mode démo, créer un fichier vide ou un placeholder
    # Note: openpyxl serait nécessaire pour créer un vrai .xlsx
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tariffs Comparison"
        ws['A1'] = "Region"
        ws['B1'] = "Average Tariff (%)"
        ws['A2'] = "AfCFTA (Intra-Africa)"
        ws['B2'] = 0.5
        ws['A3'] = "World (Extra-Africa)"
        ws['B3'] = 8.7
        wb.save(output_dir / FILES[1])
        print(f"   ✅ {FILES[1]}")
    except ImportError:
        # Si openpyxl n'est pas disponible, créer un fichier vide
        (output_dir / FILES[1]).touch()
        print(f"   ⚠️  {FILES[1]} (placeholder - openpyxl non disponible)")
    
    # 3. zlecaf_rules_of_origin.json
    rules_of_origin = {
        "generated_at": datetime.now().isoformat(),
        "rules": {
            "01-05": {"category": "Animaux vivants et produits", "rule": "Entièrement obtenus", "content": 100},
            "06-14": {"category": "Produits végétaux", "rule": "Entièrement obtenus", "content": 100},
            "25-27": {"category": "Produits minéraux", "rule": "Transformation substantielle", "content": 40},
            "28-38": {"category": "Produits chimiques", "rule": "Changement de position tarifaire", "content": 45},
            "39-40": {"category": "Plastiques et caoutchouc", "rule": "Transformation substantielle", "content": 40},
            "84-85": {"category": "Machines et équipements", "rule": "Transformation substantielle", "content": 45},
        }
    }
    
    with open(output_dir / FILES[2], 'w', encoding='utf-8') as f:
        json.dump(rules_of_origin, f, indent=2, ensure_ascii=False)
    print(f"   ✅ {FILES[2]}")
    
    # 4. zlecaf_dismantling_schedule.csv
    dismantling_schedule = [
        ["Country", "Category", "Year", "Tariff_Rate"],
        ["Algeria", "Category A", "2024", "5.0"],
        ["Algeria", "Category A", "2025", "2.5"],
        ["Algeria", "Category A", "2026", "0.0"],
        ["Egypt", "Category A", "2024", "4.0"],
        ["Egypt", "Category A", "2025", "2.0"],
        ["Egypt", "Category A", "2026", "0.0"],
        ["Kenya", "Category B", "2024", "8.0"],
        ["Kenya", "Category B", "2027", "4.0"],
        ["Kenya", "Category B", "2030", "0.0"],
    ]
    
    with open(output_dir / FILES[3], 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(dismantling_schedule)
    print(f"   ✅ {FILES[3]}")
    
    # 5. zlecaf_tariff_origin_phase.json
    tariff_phase = {
        "generated_at": datetime.now().isoformat(),
        "phases": {
            "Category A": {
                "description": "Démantèlement immédiat (0-3 ans)",
                "duration_years": 3,
                "percentage_tariff_lines": 90
            },
            "Category B": {
                "description": "Démantèlement progressif (0-5 ans)",
                "duration_years": 5,
                "percentage_tariff_lines": 7
            },
            "Category C": {
                "description": "Produits sensibles (0-10 ans)",
                "duration_years": 10,
                "percentage_tariff_lines": 3
            }
        }
    }
    
    with open(output_dir / FILES[4], 'w', encoding='utf-8') as f:
        json.dump(tariff_phase, f, indent=2, ensure_ascii=False)
    print(f"   ✅ {FILES[4]}")


def generate_production_data(output_dir: Path):
    """Génère des données de production (à implémenter avec les vraies sources)"""
    print(f"⚠️  Mode production non encore implémenté.")
    print(f"   Intégrez vos sources e-Tariff/UNCTAD/OEC ici.")
    print(f"   Utilisation du mode --demo pour le moment...")
    generate_demo_data(output_dir)


def main():
    parser = argparse.ArgumentParser(description='Génère les datasets ZLECAf')
    parser.add_argument('--demo', action='store_true', help='Générer des données de démonstration')
    parser.add_argument('--output', default='frontend/public/data', help='Répertoire de sortie')
    
    args = parser.parse_args()
    
    # Déterminer le répertoire de sortie
    if os.path.isabs(args.output):
        output_dir = Path(args.output)
    else:
        # Relative au répertoire du projet (parent du backend)
        project_root = Path(__file__).parent.parent
        output_dir = project_root / args.output
    
    output_dir.mkdir(parents=True, exist_ok=True)
    
    print(f"📦 Lyra+ Dataset Generator")
    print(f"   Répertoire de sortie: {output_dir}")
    print(f"   Mode: {'Démonstration' if args.demo else 'Production'}")
    print()
    
    if args.demo:
        generate_demo_data(output_dir)
    else:
        generate_production_data(output_dir)
    
    print()
    print(f"✅ Génération terminée - {len(FILES)} fichiers créés")
    print(f"   Fichiers disponibles dans: {output_dir}")


if __name__ == '__main__':
    main()
